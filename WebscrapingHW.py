from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from twilio.rest import Client
import keys

client = Client(keys.accountSID, keys.authToken)
TwilioNumber = "" 
mycellphone = ""

#Build Excel skeleton

wb = xl.Workbook()
ws = wb.active
ws.title = "Top Five Cryptocurrencies"

ws["A1"] = "Name"
ws["B1"] = "Symbol"
ws["C1"] = "Current Price"
ws["D1"] = "% Change in last 24hrs"
ws["E1"] = "Corresponding Price"

title_font = Font(name="Arial", size=18, color="0000cc")
table_font = Font(name="Times New Roman", size=16)

ws["A1"].font = title_font
ws["B1"].font = title_font
ws["C1"].font = title_font
ws["D1"].font = title_font
ws["E1"].font = title_font

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 36
ws.column_dimensions["E"].width = 35

#Webscrape

webpage = 'https://www.coingecko.com'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(url=webpage, headers=headers)
page = urlopen(req)
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
print(title.text)

tr = soup.findAll('tr')

for i in range(1,6):
    td = tr[i].findAll('td')

    if td:
        currency_symbol = td[2].text.split()
        name = currency_symbol[0]
        symbol = currency_symbol[1]
        price = td[3].text
        change = td[5].text.replace('%','')

        change_percent = float(change)/100
        clean_price = float(price.replace(',','').replace('$',''))
    
        if change_percent < 0:
            new_price = clean_price * (1 - change_percent)
            old_price = str("${:,.2f}".format(new_price))
        else:
            new_price = float(clean_price) * (1+ change_percent)
            old_price = str("${:,.2f}".format(new_price))

        ws['A' + str(i+1)] = name
        ws['A' + str(i+1)].font = table_font

        ws['B' + str(i+1)] = symbol
        ws['B' + str(i+1)].font = table_font

        ws['C' + str(i+1)] = price
        ws['C' + str(i+1)].font = table_font

        ws['D' + str(i+1)] = (change + '%')
        ws['D' + str(i+1)].font = table_font

        if change_percent < 0:
            new_font = Font(name="Times New Roman", size=16, color="e30000")
            ws['D' +str(i+1)].font = new_font
        else:
            new_font = Font(name="Times New Roman", size=16, color="008100")
            ws['D' +str(i+1)].font = new_font

        ws['E' + str(i+1)] = old_price
        ws['E' + str(i+1)].font = table_font

        if symbol == "BTC" or symbol == "ETH":
            
            value_check = clean_price - new_price

            increase = f'Price Alert: {symbol} increased to {price.strip()}'
            decrease = f'Price Alert: {symbol} decreased to {price.strip()}'

            if value_check > 5:
                textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                                     body=increase)
            if value_check < 5:
                textmessage = client.messages.create(to=mycellphone, from_=TwilioNumber,
                                                  body=decrease)


        wb.save("TopFiveCryptocurrencies.xlsx")